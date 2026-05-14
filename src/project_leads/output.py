from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import CrawlError, Lead


LEAD_HEADERS = [
    "日期",
    "国家",
    "项目名称",
    "项目类型",
    "业主",
    "承包商",
    "摘要",
    "来源网站",
    "原文链接",
    "相关性等级",
    "跟进建议",
]

ERROR_HEADERS = ["日期", "国家", "类别", "来源网站", "URL", "错误原因"]


def lead_to_row(lead: Lead) -> list[str]:
    return [
        lead.date,
        lead.country,
        lead.project_name,
        lead.project_type,
        lead.owner,
        lead.contractor,
        lead.summary,
        lead.source_site,
        lead.original_url,
        lead.relevance_level,
        lead.follow_up,
    ]


def error_to_row(error: CrawlError) -> list[str]:
    return [
        error.date,
        error.country,
        error.category,
        error.source_site,
        error.source_url,
        error.error_reason,
    ]


def write_excel_report(leads: list[Lead], errors: list[CrawlError], output_path: str | Path) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    leads_sheet = workbook.active
    leads_sheet.title = "Leads"
    _write_table(leads_sheet, LEAD_HEADERS, [lead_to_row(lead) for lead in leads])

    errors_sheet = workbook.create_sheet("Errors")
    _write_table(errors_sheet, ERROR_HEADERS, [error_to_row(error) for error in errors])

    workbook.save(path)
    return path


def write_markdown_report(
    leads: list[Lead],
    errors: list[CrawlError],
    output_path: str | Path,
    run_date: str,
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    high_count = sum(1 for lead in leads if lead.relevance_level == "High")
    medium_count = sum(1 for lead in leads if lead.relevance_level == "Medium")
    low_count = sum(1 for lead in leads if lead.relevance_level == "Low")

    lines = [
        "# 全球基础设施项目线索日报",
        "",
        f"生成日期：{run_date}",
        "",
        "## 汇总",
        "",
        f"- 线索总数：{len(leads)}",
        f"- 高相关：{high_count}",
        f"- 中相关：{medium_count}",
        f"- 低相关：{low_count}",
        f"- 采集失败或受限网站：{len(errors)}",
        "",
    ]

    lines.extend(_lead_sections(leads))
    lines.extend(_error_section(errors))

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def _write_table(sheet, headers: list[str], rows: list[list[str]]) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        sheet.append(row)

    sheet.freeze_panes = "A2"
    for column_index, header in enumerate(headers, start=1):
        values = [str(sheet.cell(row=row, column=column_index).value or "") for row in range(1, sheet.max_row + 1)]
        width = min(max(len(value) for value in values) + 2, 80)
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    if "原文链接" in headers:
        link_column = headers.index("原文链接") + 1
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=link_column)
            if cell.value:
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    if "URL" in headers:
        url_column = headers.index("URL") + 1
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=url_column)
            if cell.value:
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"


def _lead_sections(leads: list[Lead]) -> list[str]:
    if not leads:
        return ["## 线索列表", "", "今日未筛选到匹配线索。", ""]

    lines = ["## 线索列表", ""]
    for level in ("High", "Medium", "Low"):
        grouped = [lead for lead in leads if lead.relevance_level == level]
        if not grouped:
            continue
        lines.extend([f"### {level}", ""])
        lines.append("| 日期 | 国家 | 项目名称 | 类型 | 来源 | 原文链接 | 跟进建议 |")
        lines.append("| --- | --- | --- | --- | --- | --- | --- |")
        for lead in grouped:
            lines.append(
                "| "
                + " | ".join(
                    [
                        _md(lead.date),
                        _md(lead.country),
                        _md(lead.project_name),
                        _md(lead.project_type),
                        _md(lead.source_site),
                        f"[查看]({lead.original_url})",
                        _md(lead.follow_up),
                    ]
                )
                + " |"
            )
        lines.append("")
    return lines


def _error_section(errors: list[CrawlError]) -> list[str]:
    lines = ["## 无法直接爬取的网站", ""]
    if not errors:
        lines.extend(["无。", ""])
        return lines

    lines.append("| 日期 | 国家 | 类别 | 来源网站 | URL | 错误原因 |")
    lines.append("| --- | --- | --- | --- | --- | --- |")
    for error in errors:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(error.date),
                    _md(error.country),
                    _md(error.category),
                    _md(error.source_site),
                    f"[打开]({error.source_url})",
                    _md(error.error_reason),
                ]
            )
            + " |"
        )
    lines.append("")
    return lines


def _md(value: str) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", " ")
